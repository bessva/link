# -*- coding: utf-8 -*-
"""
линк_v5_app.py — Streamlit интерфейс для агента + калькулятор нормативов
Запуск: streamlit run линк_v5_app.py
"""

import os
import re
import uuid
import time
import glob
import requests
import streamlit as st
from datetime import datetime

# ==============================
# ⚙️ НАСТРОЙКИ
# ==============================

GIGACHAT_AUTH_KEY = "MDE5Y2E1MTQtMzEzYS03NzRkLTk0NWEtYTk0YjU0ZGFlMmM3OmE2ODk0ZjViLTMyMmMtNDUwYy1hZTdiLTRlY2ZiOGMxMTdlNw=="

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = os.path.join(BASE_DIR, "data")
KNOWLEDGE_DIR = os.path.join(BASE_DIR, "knowledge")
VECTOR_DB_DIR = os.path.join(BASE_DIR, "vector_db")

for d in [DATA_DIR, KNOWLEDGE_DIR, VECTOR_DB_DIR]:
    os.makedirs(d, exist_ok=True)

# ==============================
# 🎨 Стили
# ==============================

st.set_page_config(
    page_title="ЛИНК — Энергетический ассистент",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.google.com/specimen/Montserrat?lang=ru_Cyrl');

.stApp {
    background-color: #0d1117;
    font-family: 'Montserrat', sans-serif;
}
[data-testid="stSidebar"] {
    background-color: #161b22;
    border-right: 1px solid #21262d;
}
[data-testid="stSidebar"] * { color: #c9d1d9 !important; }
.sidebar-title {
    font-family: 'Montserrat', monospace;
    font-size: 18px; font-weight: 600; color: #58a6ff !important;
    letter-spacing: 2px; text-transform: uppercase; margin-bottom: 4px;
}
.sidebar-subtitle {
    font-size: 11px; color: #8b949e !important;
    letter-spacing: 1px; text-transform: uppercase; margin-bottom: 24px;
}
.sidebar-section {
    font-size: 10px; font-weight: 600; letter-spacing: 2px;
    text-transform: uppercase; color: #58a6ff !important;
    margin: 20px 0 8px 0; padding-bottom: 4px; border-bottom: 1px solid #21262d;
}
.sidebar-item {
    font-size: 12px; color: #8b949e !important;
    padding: 3px 0; display: flex; align-items: center; gap: 6px;
}
.sidebar-item-active { color: #c9d1d9 !important; }
.main-header {
    font-family: 'Montserrat', monospace; font-size: 13px; color: #58a6ff;
    letter-spacing: 3px; text-transform: uppercase;
    padding: 16px 0 4px 0; border-bottom: 1px solid #21262d; margin-bottom: 24px;
}
.chat-container { display: flex; flex-direction: column; gap: 16px; padding-bottom: 20px; }
.msg-user {
    background: #1c2128; border: 1px solid #30363d;
    border-radius: 8px 8px 2px 8px; padding: 12px 16px;
    margin-left: 60px; color: #c9d1d9; font-size: 14px; line-height: 1.6;
}
.msg-assistant {
    background: #161b22; border: 1px solid #21262d;
    border-left: 3px solid #58a6ff; border-radius: 2px 8px 8px 8px;
    padding: 12px 16px; margin-right: 60px; color: #c9d1d9;
    font-size: 14px; line-height: 1.6;
}
.msg-label-user {
    font-size: 10px; color: #8b949e; letter-spacing: 1px;
    text-transform: uppercase; margin-bottom: 6px; text-align: right;
}
.msg-label-assistant {
    font-size: 10px; color: #58a6ff; letter-spacing: 1px;
    text-transform: uppercase; margin-bottom: 6px; font-family: 'Montserrat', monospace;
}
.source-badge {
    display: inline-block; font-size: 9px; letter-spacing: 1px;
    text-transform: uppercase; padding: 2px 6px; border-radius: 3px;
    margin-right: 4px; font-family: 'Montserrat', monospace;
}
.source-excel { background: #1a3a1a; color: #3fb950; border: 1px solid #238636; }
.source-kb    { background: #1a1f3a; color: #79c0ff; border: 1px solid #1f6feb; }
.source-both  { background: #2d1f3a; color: #d2a8ff; border: 1px solid #6e40c9; }
.source-calc  { background: #2d2a1a; color: #f0c040; border: 1px solid #9e7c00; }
.stTextInput input {
    background-color: #161b22 !important; border: 1px solid #30363d !important;
    border-radius: 6px !important; color: #c9d1d9 !important;
    font-family: 'Montserrat', sans-serif !important;
    font-size: 14px !important; padding: 10px 14px !important;
}
.stTextInput input:focus {
    border-color: #58a6ff !important;
    box-shadow: 0 0 0 3px rgba(88, 166, 255, 0.1) !important;
}
.stButton button {
    background-color: #21262d !important; color: #c9d1d9 !important;
    border: 1px solid #30363d !important; border-radius: 6px !important;
    font-family: 'Montserrat', sans-serif !important;
    font-size: 12px !important; letter-spacing: 0.5px !important; transition: all 0.2s !important;
}
.stButton button:hover {
    background-color: #30363d !important; border-color: #58a6ff !important; color: #58a6ff !important;
}
.status-ok  { color: #3fb950; font-size: 11px; }
.status-err { color: #f85149; font-size: 11px; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 20px !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500;600&display=swap" rel="stylesheet">
""", unsafe_allow_html=True)


# ==============================
# 🔐 GigaChat
# ==============================

def get_access_token():
    if ("_token" in st.session_state and
            time.time() < st.session_state.get("_token_exp", 0) - 60):
        return st.session_state["_token"]

    url = "https://ngw.devices.sberbank.ru:9443/api/v2/oauth"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
        "RqUID": str(uuid.uuid4()),
        "Authorization": f"Basic {GIGACHAT_AUTH_KEY}",
    }
    r = requests.post(url, headers=headers, data={"scope": "GIGACHAT_API_PERS"},
                      verify=False, timeout=15)
    r.raise_for_status()
    d = r.json()
    st.session_state["_token"]     = d["access_token"]
    st.session_state["_token_exp"] = d.get("expires_at", 0) / 1000
    return st.session_state["_token"]


def call_gigachat(system_prompt, user_message, history):
    token = get_access_token()
    messages = [{"role": "system", "content": system_prompt}]
    for q, a, _ in history[-5:]:
        messages.append({"role": "user",      "content": q})
        messages.append({"role": "assistant", "content": a})
    messages.append({"role": "user", "content": user_message})

    r = requests.post(
        "https://gigachat.devices.sberbank.ru/api/v1/chat/completions",
        headers={"Accept": "application/json", "Content-Type": "application/json",
                 "Authorization": f"Bearer {token}"},
        json={"model": "GigaChat-Pro", "messages": messages,
              "temperature": 0.3, "max_tokens": 1000},
        verify=False, timeout=30
    )
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"].strip()


# ==============================
# 📊 Excel
# ==============================

def _find_col(headers, keywords):
    for i, h in enumerate(headers):
        if h and any(kw in str(h).lower() for kw in keywords):
            return i
    return None

@st.cache_data(show_spinner=False)
def load_excel_data(folder):
    try:
        import openpyxl
    except ImportError:
        return {}
    all_data = {}
    for filepath in glob.glob(os.path.join(folder, "*.xlsx")):
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            filename = os.path.basename(filepath)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                # Заголовки — обрезаем до 30 символов
                headers = [str(h)[:80] if h else f"col_{i}" for i, h in enumerate(rows[0])]

                # Читаем данные с forward-fill по первым 4 колонкам (название, топливо и т.д.)
                prev = [None] * len(headers)
                lines = [" | ".join(headers)]

                for row in rows[2:]:  # пропускаем заголовок и пустую строку
                    filled = []
                    for i, c in enumerate(row):
                        if c is not None:
                            prev[i] = c
                            filled.append(str(c))
                        elif i < 4 and prev[i] is not None:
                            filled.append(str(prev[i]))
                        else:
                            filled.append("")
                    # Пропускаем полностью пустые строки
                    if any(v for v in filled):
                        lines.append(" | ".join(filled))

                key = f"{filename}___{sheet_name}"
                all_data[key] = "\n".join(lines)
        except Exception:
            pass
    return all_data


def find_relevant_excel(excel_data, question):
    q = question.lower()
    result_parts = []
    for sheet_key, text in excel_data.items():
        words = [w for w in q.split() if len(w) > 3]
        if any(w[:5] in sheet_key.lower() or w[:5] in text[:300].lower() for w in words):
            result_parts.append(f"=== {sheet_key} ===\n{text}")
    if not result_parts:
        result_parts = [f"=== {k} ===\n{v}" for k, v in excel_data.items()]
    return "\n\n".join(result_parts)


# ==============================
# 📚 База знаний
# ==============================

def extract_pdf(path):
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t: text += t + "\n"
        return text
    except Exception:
        return ""

def extract_docx(path):
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception:
        return ""

def chunk_text(text, size=800, overlap=100):
    words = text.split()
    chunks, i = [], 0
    while i < len(words):
        c = " ".join(words[i:i+size])
        if c.strip(): chunks.append(c)
        i += size - overlap
    return chunks

@st.cache_resource(show_spinner=False)
def build_knowledge_base(knowledge_dir, vector_db_dir):
    try:
        import chromadb
        from sentence_transformers import SentenceTransformer
    except ImportError:
        return None, None

    files = (glob.glob(os.path.join(knowledge_dir, "*.pdf")) +
             glob.glob(os.path.join(knowledge_dir, "*.docx")))
    if not files:
        return None, None

    client = chromadb.PersistentClient(path=vector_db_dir)
    existing = [c.name for c in client.list_collections()]
    model = SentenceTransformer("paraphrase-multilingual-mpnet-base-v2")

    if "knowledge" in existing:
        col = client.get_collection("knowledge")
        if col.count() > 0:
            return col, model

    if "knowledge" in existing:
        client.delete_collection("knowledge")
    col = client.create_collection("knowledge")

    all_chunks, all_ids, all_metas = [], [], []
    for filepath in files:
        fname = os.path.basename(filepath)
        ext   = os.path.splitext(filepath)[1].lower()
        text  = extract_pdf(filepath) if ext == ".pdf" else extract_docx(filepath)
        if not text.strip(): continue
        for i, chunk in enumerate(chunk_text(text)):
            all_chunks.append(chunk)
            all_ids.append(f"{fname}_{i}")
            all_metas.append({"source": fname, "chunk": i})

    if all_chunks:
        embeddings = model.encode(all_chunks, show_progress_bar=False).tolist()
        col.add(documents=all_chunks, embeddings=embeddings,
                ids=all_ids, metadatas=all_metas)
    return col, model


def search_kb(question, collection, model, top_k=3):
    if collection is None or model is None: return ""
    try:
        q_emb   = model.encode([question]).tolist()
        results = collection.query(query_embeddings=q_emb, n_results=top_k)
        docs    = results.get("documents", [[]])[0]
        metas   = results.get("metadatas", [[]])[0]
        return "\n\n---\n\n".join(f"[{m.get('source','?')}]\n{d}" for d, m in zip(docs, metas))
    except Exception:
        return ""


# ==============================
# 🧮 Калькулятор нормативов
# ==============================

PROMPTS_DB = {
    "nnzt_piped": {
        "keywords": ["неснижаем", "ннзт", "труба", "трубопровод"],
        "formula": "3 × В_усл × (7000 / Q_р)",
        "calc_function": lambda V_usl, Q_pr: 3 * V_usl * (7000 / Q_pr),
        "params": ["V_usl", "Q_pr"],
        "param_prompts": {
            "V_usl": "суточный расход условного топлива в режиме выживания (т.у.т.)",
            "Q_pr":  "низшая теплота сгорания натурального топлива (ккал/кг)"
        }
    },
    "nnzt_general": {
        "keywords": ["неснижаем", "ннзт", "уголь", "мазут"],
        "formula": "7 × В_усл × (7000 / Q_р)",
        "calc_function": lambda V_usl, Q_pr: 7 * V_usl * (7000 / Q_pr),
        "params": ["V_usl", "Q_pr"],
        "param_prompts": {
            "V_usl": "суточный расход условного топлива в режиме выживания (т.у.т.)",
            "Q_pr":  "низшая теплота сгорания натурального топлива (ккал/кг)"
        }
    },
    "nezt_small": {
        "keywords": ["нэзт", "нормальный эксплуатационный", "менее 25", "<25"],
        "formula": "В_ср × Т_кор × К_ср",
        "calc_function": lambda V_sr, T_kor, K_sr: V_sr * T_kor * K_sr,
        "params": ["V_sr", "T_kor", "K_sr"],
        "param_prompts": {
            "V_sr":  "среднесуточный расход натурального топлива (т.н.т.)",
            "T_kor": "корректирующий период (сутки)",
            "K_sr":  "коэффициент срыва поставок (1.2–3.5)"
        }
    },
    "nezt_large_piped": {
        "keywords": ["нэзт", "нормальный эксплуатационный", "газ", "более 25", "≥25"],
        "formula": "НЭЗТ_б.в. × R_тэс",
        "calc_function": lambda NEZT_bv, R_tes: NEZT_bv * R_tes,
        "params": ["NEZT_bv", "R_tes"],
        "param_prompts": {
            "NEZT_bv": "базовая величина НЭЗТ (т.н.т.)",
            "R_tes":   "коэффициент риска R_ТЭС (0–1)"
        }
    },
    "nezt_large_solid": {
        "keywords": ["нэзт", "нормальный эксплуатационный", "уголь", "торф", "твердое"],
        "formula": "НЭЗТ_б.в. × К_пост × К_ср",
        "calc_function": lambda NEZT_bv, K_post, K_sr: NEZT_bv * K_post * K_sr,
        "params": ["NEZT_bv", "K_post", "K_sr"],
        "param_prompts": {
            "NEZT_bv": "базовая величина НЭЗТ (т.н.т.)",
            "K_post":  "коэффициент времени поставки (0.3–1.8)",
            "K_sr":    "коэффициент срыва поставки (1.0–2.5)"
        }
    },
    "nazt_small": {
        "keywords": ["назт", "нормативный аварийный", "менее 25", "аварийное"],
        "formula": "В_сут × N × k / 24",
        "calc_function": lambda V_sut, N, k: V_sut * N * k / 24,
        "params": ["V_sut", "N", "k"],
        "param_prompts": {
            "V_sut": "максимальный суточный расход аварийного топлива (т.н.т.)",
            "N":     "количество суток работы на аварийном топливе (3–5)",
            "k":     "количество часов работы в сутки"
        }
    },
    "nazt_large": {
        "keywords": ["назт", "нормативный аварийный", "более 25", "≥25"],
        "formula": "3 × В_сут × R_тэс",
        "calc_function": lambda V_sut, R_tes: 3 * V_sut * R_tes,
        "params": ["V_sut", "R_tes"],
        "param_prompts": {
            "V_sut": "максимальный суточный расход аварийного топлива (т.н.т.)",
            "R_tes": "коэффициент риска R_ТЭС (0–1)"
        }
    },
    "nvzt_large": {
        "keywords": ["нвзт", "нормативный вспомогательный", "растопк", "вспомогательн"],
        "formula": "ОНЗТ × (В_всп / В_осн) + V_р/о + V_ав",
        "calc_function": lambda ONZT, V_vsp, V_osn, V_ro, V_av: ONZT * (V_vsp / V_osn) + V_ro + V_av,
        "params": ["ONZT", "V_vsp", "V_osn", "V_ro", "V_av"],
        "param_prompts": {
            "ONZT":  "ОНЗТ (т.н.т.)",
            "V_vsp": "фактический расход вспомогательного топлива за 3 года (т.н.т.)",
            "V_osn": "фактический расход основного топлива за 3 года (т.н.т.)",
            "V_ro":  "запас для растопок V_р/о (т.н.т.)",
            "V_av":  "запас для ликвидации аварий V_ав (т.н.т.)"
        }
    }
}

CALC_KEYWORDS = ["рассчитай", "посчитай", "вычисли", "расчёт", "расчет", "калькулятор",
                 "формула", "по формуле", "сколько будет"]

def select_calc_prompt(q):
    q_lower = q.lower()
    # Срабатывает только если есть явный запрос на расчёт
    if not any(kw in q_lower for kw in CALC_KEYWORDS):
        return None, None
    priority = ["nvzt_large", "nazt_large", "nazt_small",
                "nnzt_piped", "nnzt_general",
                "nezt_large_piped", "nezt_large_solid", "nezt_small"]
    for key in priority:
        if any(kw in q_lower for kw in PROMPTS_DB[key]["keywords"]):
            return key, PROMPTS_DB[key]
    return None, None


# ==============================
# 🧠 Классификация вопроса
# ==============================

def classify_question(q):
    return "both"


SYSTEM_PROMPT = (
    "Ты — эксперт по управлению энергетическими объектами и нормативам "
    "топливных запасов (ННЗТ, НЭЗТ, НАЗТ, ОНЗТ, НВЗТ). "
    "Ты умеешь анализировать числовые данные по станциям и давать "
    "управленческие рекомендации на основе системно-ситуационного подхода. "
    "Отвечай КРАТКО и по существу — максимум 3-5 предложений или короткий список. "
    "Без вступлений, без повторений вопроса, без лишних слов. Только факты и выводы. "
    "ВАЖНО: отвечай ТОЛЬКО на вопросы про энергетику, ТЭЦ, ГРЭС, топливные запасы "
    "и управление энергообъектами. На остальное отвечай: "
    "'Я специализируюсь на энергетике и не отвечаю на вопросы вне этой темы.'"
)


# ==============================
# 🖥️ Интерфейс
# ==============================

if "history" not in st.session_state:
    st.session_state.history = []
if "calc_state" not in st.session_state:
    st.session_state.calc_state = None  # {key, data, collected_params}

excel_data = load_excel_data(DATA_DIR)

with st.spinner("Загружаю базу знаний..."):
    collection, st_model = build_knowledge_base(KNOWLEDGE_DIR, VECTOR_DB_DIR)

# ── Сайдбар ──────────────────────────────────────────

with st.sidebar:
    st.markdown('<div class="sidebar-title">⚡ ЛИНК</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-subtitle">Энергетический ассистент</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-section">Станции</div>', unsafe_allow_html=True)
    if excel_data:
    # Показываем только уникальные имена файлов без листов
        seen = set()
        for key in excel_data.keys():
            filename = key.split("___")[0].replace(".xlsx", "")
            if filename not in seen:
                seen.add(filename)
                st.markdown(f'<div class="sidebar-item sidebar-item-active">🏭 {filename}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="sidebar-item">Нет данных в папке data/</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-section">База знаний</div>', unsafe_allow_html=True)
    kb_files = (glob.glob(os.path.join(KNOWLEDGE_DIR, "*.pdf")) +
                glob.glob(os.path.join(KNOWLEDGE_DIR, "*.docx")))
    if kb_files:
        for f in kb_files:
            st.markdown(f'<div class="sidebar-item sidebar-item-active">📄 {os.path.basename(f)}</div>',
                        unsafe_allow_html=True)
    else:
        st.markdown('<div class="sidebar-item">Нет документов в папке knowledge/</div>',
                    unsafe_allow_html=True)

    st.markdown('<div class="sidebar-section">Калькулятор</div>', unsafe_allow_html=True)
    for key, data in PROMPTS_DB.items():
        st.markdown(f'<div class="sidebar-item">📐 {data["formula"]}</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("🗑  Очистить чат"):
        st.session_state.history = []
        st.session_state.calc_state = None
        st.rerun()

    st.markdown('<div class="sidebar-section">Статус</div>', unsafe_allow_html=True)
    try:
        get_access_token()
        st.markdown('<div class="status-ok">● GigaChat подключён</div>', unsafe_allow_html=True)
    except Exception:
        st.markdown('<div class="status-err">● Ошибка подключения</div>', unsafe_allow_html=True)

    if collection is not None:
        st.markdown(f'<div class="status-ok">● База знаний активна ({collection.count()} фрагм.)</div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div class="status-err">● База знаний не загружена</div>', unsafe_allow_html=True)


# ── Основная область ─────────────────────────────────

st.markdown('<div class="main-header">// Диалог с ассистентом</div>', unsafe_allow_html=True)

for user_msg, bot_msg, source in st.session_state.history:
    source_badge = {
        "excel":     '<span class="source-badge source-excel">📊 Данные</span>',
        "knowledge": '<span class="source-badge source-kb">📚 База знаний</span>',
        "both":      '<span class="source-badge source-both">📊📚 Оба источника</span>',
        "calc":      '<span class="source-badge source-calc">📐 Калькулятор</span>',
    }.get(source, "")

    st.markdown(f"""
        <div class="msg-label-user">Вы</div>
        <div class="msg-user">{user_msg}</div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
        <div class="msg-label-assistant">⚡ ЛИНК {source_badge}</div>
        <div class="msg-assistant">{bot_msg}</div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
with st.form("chat_form", clear_on_submit=True):
    col1, col2 = st.columns([6, 1])
    with col1:
        user_input = st.text_input(
            "",
            placeholder="Вопрос или 'рассчитай ННЗТ' / 'посчитай НАЗТ'...",
            label_visibility="collapsed"
        )
    with col2:
        submitted = st.form_submit_button("Отправить", use_container_width=True)

if submitted and user_input.strip():

    # ── Режим сбора параметров калькулятора ──────────
    if st.session_state.calc_state is not None:
        cs = st.session_state.calc_state
        param_name = cs["data"]["params"][len(cs["collected_params"])]

        # Попытка распарсить число
        nums = re.findall(r"[-+]?\d*\.?\d+", user_input)
        if nums:
            cs["collected_params"][param_name] = float(nums[0])
            remaining = [p for p in cs["data"]["params"] if p not in cs["collected_params"]]

            if not remaining:
                # Все параметры собраны — считаем
                try:
                    result = cs["data"]["calc_function"](**cs["collected_params"])
                    answer = (f"📐 Формула: {cs['data']['formula']}\n"
                              f"Параметры: {cs['collected_params']}\n\n"
                              f"✅ Результат: {result:.3f} т.н.т.")
                except Exception as e:
                    answer = f"❌ Ошибка расчёта: {e}"
                st.session_state.calc_state = None
                st.session_state.history.append((user_input, answer, "calc"))
            else:
                # Запрашиваем следующий параметр
                next_param = remaining[0]
                answer = f"Принято. Теперь введите: {cs['data']['param_prompts'][next_param]}"
                st.session_state.history.append((user_input, answer, "calc"))
        else:
            answer = "⚠️ Не распознано число. Введите числовое значение."
            st.session_state.history.append((user_input, answer, "calc"))

        st.rerun()

    # ── Обычный режим ─────────────────────────────────
    else:
        calc_key, calc_data = select_calc_prompt(user_input)

        if calc_key:
            # Пробуем вытащить числа прямо из вопроса
            nums = [float(n) for n in re.findall(r"[-+]?\d*\.?\d+", user_input)]
            params_needed = calc_data["params"]

            if len(nums) >= len(params_needed):
                params = dict(zip(params_needed, nums[:len(params_needed)]))
                try:
                    result = calc_data["calc_function"](**params)
                    answer = (f"📐 Формула: {calc_data['formula']}\n"
                              f"Параметры: {params}\n\n"
                              f"✅ Результат: {result:.3f} т.н.т.")
                except Exception as e:
                    answer = f"❌ Ошибка расчёта: {e}"
                st.session_state.history.append((user_input, answer, "calc"))
            else:
                # Начинаем пошаговый сбор параметров
                first_param = params_needed[0]
                st.session_state.calc_state = {
                    "key": calc_key,
                    "data": calc_data,
                    "collected_params": {}
                }
                answer = (f"📐 Запускаю расчёт по формуле: {calc_data['formula']}\n\n"
                          f"Введите: {calc_data['param_prompts'][first_param]}")
                st.session_state.history.append((user_input, answer, "calc"))

        else:
            # Обычный вопрос — Excel + база знаний + GigaChat
            q_type = classify_question(user_input)
            context_parts = []

            if excel_data:
                ec = find_relevant_excel(excel_data, user_input)
                if ec: context_parts.append(f"📊 ДАННЫЕ ПО СТАНЦИЯМ:\n{ec}")

            if collection is not None:
                kc = search_kb(user_input, collection, st_model)
                if kc: context_parts.append(f"📚 ИЗ БАЗЫ ЗНАНИЙ:\n{kc}")

            user_message = user_input + ("\n\n" + "\n\n".join(context_parts) if context_parts else "")

            with st.spinner("Думаю..."):
                try:
                    answer = call_gigachat(SYSTEM_PROMPT, user_message, st.session_state.history)
                    answer = re.sub(r'\*+', '', answer)
                except Exception as e:
                    answer = f"❌ Ошибка: {e}"

            st.session_state.history.append((user_input, answer, q_type))

        st.rerun()
